import os
import cv2
import numpy as np
import math
import time
from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = "Coordinates"
ws.append(["X", "Y"])  # 表頭: X 和 Y
excel_file = "coordinates.xlsx"
folder_name = "test_path"  # 定義文件夾名稱
# 定義 FLANN 參數
FLANN_INDEX_KDTREE = 1
index_params = dict(algorithm=FLANN_INDEX_KDTREE, trees=5)
search_params = dict(checks=50)
flann = cv2.FlannBasedMatcher(index_params, search_params)  # 創建 FLANN 匹配器
# 初始化 SIFT 檢測器
sift = cv2.SIFT_create()
def split_image(image, num_rows, num_cols):
    # 獲取圖像的高度和寬度
    height, width = image.shape[:2]
    print(f"height, width:{height}, {width}")
    # 計算每個區塊的高度和寬度
    block_height = height // num_rows
    block_width = width // num_cols
    print(f"block_height, block_width:{block_height}, {block_width}")
    # 存儲所有區塊的列表
    blocks = []
    for r in range(num_rows):
        for c in range(num_cols):
            # 計算當前區塊的範圍
            start_row = r * block_height
            end_row = (r + 1) * block_height
            start_col = c * block_width
            end_col = (c + 1) * block_width
            # 提取當前區塊
            block = image[start_row:end_row, start_col:end_col]
            blocks.append(block)
    return blocks,block_height,block_width
def clear_excel_content(file_name):
    try:
        # 加載現有的 Excel 檔案
        wb = load_workbook(file_name)
        for sheet in wb.worksheets:
            # 清空每個工作表內容
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None
        # 儲存更改
        wb.save(file_name)
        print(f"Excel 檔案 {file_name} 已清空內容！")
    except FileNotFoundError:
        print(f"檔案 {file_name} 不存在！")

def match_features(query_image, train_images):
    # 初始化 SIFT 檢測器
    sift = cv2.SIFT_create()
    # 使用 SIFT 找到 query_image 的關鍵點和描述子
    kp1, des1 = sift.detectAndCompute(query_image, None)
    # 創建 FLANN 匹配器
    flann = cv2.FlannBasedMatcher()
    # 儲存所有匹配的結果
    all_matches = []
    for train_image in train_images:
        # 使用 SIFT 找到 train_image 的關鍵點和描述子
        kp2, des2 = sift.detectAndCompute(train_image, None)
        # 進行特徵匹配
        matches = flann.knnMatch(des1, des2, k=2)
        # 剔除不好的匹配
        good_matches = []
        for m, n in matches:
            if m.distance < 0.7 * n.distance:
                good_matches.append(m)
        all_matches.append(len(good_matches))
        
    return all_matches

def find_most_matched_block(matches):
    most_matched_block_index = np.argmax(matches)
    return most_matched_block_index
def find_surrounding_indices(most_matched_block_index, num_rows, num_cols):
    # most_matched_block_index=15
    row_index = most_matched_block_index // num_cols
    col_index = most_matched_block_index % num_cols

    surrounding_indices = []
    offsets = [(-1, -1), (-1, 0), (-1, 1), (0, -1), (0, 1), (1, -1), (1, 0), (1, 1)]

    for dr, dc in offsets:
        r = row_index + dr
        c = col_index + dc
        if 0 <= r < num_rows and 0 <= c < num_cols:
            surrounding_indices.append(r * num_cols + c)

    return surrounding_indices
    # for dr in range(-1, 2):
    #     for dc in range(-1, 2):
    #         r = row_index + dr
    #         c = col_index + dc
    #         if (dr != 0 or dc != 0) and 0 <= r < num_rows and 0 <= c < num_cols:
    #             surrounding_indices.append(r * num_cols + c)
    #
    # return surrounding_indices

# 定義函數來將索引與對應的區塊匹配並返回結果列表


def show_matched_blocks(query_image, matched_indices, blocks):
    for index in matched_indices:
        if 0 <= index < len(blocks):
            block = blocks[index]
            cv2.imshow(f'Matched Block {index}', block)
    cv2.waitKey(0)
    cv2.destroyAllWindows()
###不知道為何合併找最佳索引時會出錯，先留著
def find_most_matched_and_surrounding_indices(matches, all_blocks, num_rows, num_cols):
    most_matched_block_index = np.argmax(matches)

    best_index = None
    for j in range(num_cols * num_rows):
        if np.array_equal(all_blocks[j], all_blocks[most_matched_block_index]):
            best_index = j
            break

    surrounding_indices = find_surrounding_indices(best_index, num_rows, num_cols)

    return best_index, surrounding_indices

def merge_blocks_into_one_image(blocks, best_index, num_rows, num_cols):
    block_height, block_width = all_blocks[0].shape[:2]

    if len(blocks) == 9:  # 3x3
        merged_image = np.zeros((3 * block_height // 2, 3 * block_width // 2, 3), dtype=np.uint8)
        rows, cols = 3, 3
    elif len(blocks) == 6:  # 2x3
        merged_image = np.zeros((2 * block_height // 2, 3 * block_width // 2, 3), dtype=np.uint8)
        rows, cols = 2, 3
    elif len(blocks) == 4:  # 2x2
        merged_image = np.zeros((2 * block_height // 2, 2 * block_width // 2, 3), dtype=np.uint8)
        rows, cols = 2, 2
    else:
        raise ValueError("不支援的塊數量")

    # 確保所有塊的尺寸一致
    target_height = block_height // 2
    target_width = block_width // 2

    for r in range(rows):
        for c in range(cols):
            block_index = r * cols + c
            if block_index < len(blocks):
                block = blocks[block_index]
                # 調整圖像塊的大小以確保尺寸一致
                block_resized = cv2.resize(block, (target_width, target_height))
                merged_image[r * target_height:(r + 1) * target_height, c * target_width:(c + 1) * target_width] = block_resized

    return merged_image


# def merge_blocks_into_one_image(blocks, num_rows, num_cols):
#     block_height, block_width = all_blocks[0].shape[:2]
#     merged_image = np.zeros((num_rows * block_height // 2, num_cols * block_width // 2, 3), dtype=np.uint8)
#     if len(blocks) == 9 :
#         num_rows = 3
#         num_cols = 3
#         for r in range(num_rows):
#             for c in range(num_cols):
#                 block_index = r * num_cols + c
#                 block = cv2.resize(blocks[block_index], (block_width // 2, block_height // 2))  # Resize each block
#                 merged_image[r * block_height // 2:(r + 1) * block_height // 2, c * block_width // 2:(c + 1) * block_width // 2] = block
#     elif len(blocks) == 4 :
#         num_rows = 2
#         num_cols = 2
#     return merged_image

def rotate_point(x, y, cx, cy, angle):
    # Convert angle to radians
    theta = math.radians(angle)
    # Perform rotation
    x_rotated = cx + (x - cx) * math.cos(theta) - (y - cy) * math.sin(theta)
    y_rotated = cy + (x - cx) * math.sin(theta) + (y - cy) * math.cos(theta)
    return x_rotated, y_rotated
def identify(center_points,merge_image,img,best_index,num_rows,num_cols,block_width,block_height):
    kp, des = sift.detectAndCompute(merge_image, None)
    kp1, des1 = sift.detectAndCompute(img, None)

    # 進行特徵匹配
    matches1 = flann.knnMatch(des1, des, k=2)
    # 根據 Lowe's ratio 測試存儲所有良好的匹配
    good1 = []
    for m1, n1 in matches1:
        if m1.distance < 0.7 * n1.distance:
            good1.append(m1)
    #homography
    if len(good1) > 10:
        src_pts = []
        dst_pts = []
        for m1 in good1:
            src_pts.append(kp1[m1.queryIdx].pt)
            dst_pts.append(kp[m1.trainIdx].pt)

        src_pts = np.float32(src_pts).reshape(-1, 1, 2)
        dst_pts = np.float32(dst_pts).reshape(-1, 1, 2)
        M1, mask = cv2.findHomography(src_pts, dst_pts, cv2.RANSAC, 5.0)


    center_x = np.mean([pt[0][0] for pt in dst_pts])
    center_y = np.mean([pt[0][1] for pt in dst_pts])

    # 使用透視變換將merge地圖的中心座標映射到big地圖上
    center = np.float32([[center_x, center_y]])
    center = np.array([center], dtype=np.float32)
    merge_map_center = cv2.perspectiveTransform(center, M1)

    # 得到在big_map_image上的中心座標
    big_map_center_x = merge_map_center[0][0][0]+(((best_index-1) %  num_cols) +0.5) * block_width
    big_map_center_y = merge_map_center[0][0][1]+(((best_index-1) // num_cols) ) * block_height

    ws.append([big_map_center_x,big_map_center_y])
    wb.save(excel_file)
    print(f"數據已成功儲存至 {excel_file}")
    center_points.append((big_map_center_x, big_map_center_y))
    print(merge_map_center[0][0][0],merge_map_center[0][0][1])
    # print(center_points)
    rotation_angle = np.arctan2(M1[1, 0], M1[0, 0]) * 180 / np.pi
    return center_points ,rotation_angle

def get_blocks_by_indices(all_blocks, indices):
    """
    根據索引號碼從所有區塊中取得對應的區塊陣列。

    參數：
        all_blocks (list): 包含所有區塊的列表。
        indices (list): 索引號碼的列表，用於指示要提取的區塊。

    返回：
        list: 包含對應索引的區塊陣列。
    """
    blocks = [all_blocks[index] for index in indices]
    return blocks

big_map_img = cv2.imread("Big_map_collect/2024-12-25-bigmap.jpg")
# 指定要分割的行和列數
num_rows = 6
num_cols = 6

new_num_rows = 3
new_num_cols = 3
all_blocks,block_height,block_width = split_image(big_map_img, num_rows, num_cols)
blocks = all_blocks
# output_folder_name = "output_frames"
# num = len(os.listdir("output_frames"))
output_folder_name = "captured_frames"
num = len(os.listdir("captured_frames"))
center_points = []
# -------------------------------------------------------------------------
def main(blocks):
    total_path = 0

    for i in range(0, num+1):
        image = cv2.imread(f'{output_folder_name}/frame_{i}.jpg')
        matches = match_features(image, blocks)
        # #測試
        # best_index, surrounding_indices = find_most_matched_and_surrounding_indices(matches, all_blocks, num_rows,
        #                                                                             num_cols)
        #
        # print(f'最適合區域索引 {best_index}')
        # print(f'surrounding_indices = {surrounding_indices}')
        #
        # most_matched_block = all_blocks[best_index]
        # surrounding_blocks = [all_blocks[i] for i in surrounding_indices]
        # # 將最匹配的區塊和周圍的區塊合併為一個列表
        # blocks_to_match = [most_matched_block] + surrounding_blocks
        # block_indices = []
        most_matched_block_index = find_most_matched_block(matches)
        best_index = None
        for j in range(0, (num_cols * num_rows)):
            if np.array_equal(all_blocks[j], blocks[most_matched_block_index]):
                best_index = j
                break
        print(f'最適合區域索引{best_index}')
        surrounding_indices = find_surrounding_indices(best_index, num_rows, num_cols)

        print(surrounding_indices)
        # 取得最匹配的區塊以及其周圍的區塊
        most_matched_block = all_blocks[best_index]
        surrounding_blocks = [all_blocks[i] for i in surrounding_indices]
        # 將最匹配的區塊和周圍的區塊合併為一個列表
        blocks_to_match = [most_matched_block] + surrounding_blocks
        block_indices = []

        for i, block_to_match in enumerate(blocks_to_match):
            # 初始化索引
            block_index = None
            # 尋找 block_to_match 在 blocks 中的索引
            for j, block in enumerate(all_blocks):
                # 如果兩個陣列相等，則找到索引
                if np.array_equal(block_to_match, block):
                    block_index = j
                    break

            if block_index is not None:
                block_indices.append(block_index)
                # print(f"區塊 {i}: 在 blocks 中的索引 = {block_index}")
            else:
                print(f"區塊 {i}: 找不到在 blocks 中的對應索引")
        block_indices_sorted = sorted(block_indices)
        print(block_indices_sorted)
        # show_matched_blocks(image, block_indices_sorted, blocks)
        # merged_image = merge_blocks_into_one_image([all_blocks[index] for index in block_indices_sorted], 3, 3)
        merged_image = merge_blocks_into_one_image([all_blocks[index] for index in block_indices_sorted], best_index, num_rows, num_cols)
        # center_points = []

        center, rotation_angle = identify(center_points,merged_image, image,best_index,num_rows,num_cols,block_width,block_height)


        for point in center_points:
            center_x, center_y = map(int, point)
            cv2.circle(big_map_img, (center_x, center_y), radius=5, color=(255, 255, 255), thickness=25)
        base_length = 50
        half_base = base_length // 2
        height = int(0.866 * base_length)  # Height of an equilateral triangle is sqrt(3)/2 times the base length
        offset_factor = 50  # Adjust this value as needed
        # offset = offset_factor * angle / abs(angle) if angle != 0 else offset_factor

        # Calculate coordinates of base vertices
        vertex1 = (int(center_x - half_base - offset_factor), int(center_y + height))
        vertex2 = (int(center_x - offset_factor), int(center_y - height))
        vertex3 = (int(center_x + half_base - offset_factor), int(center_y + height))

        # Rotate the base vertices around the center point by a given angle (e.g., 45 degrees)
        # rotation_angle = 45  # Specify the rotation angle here
        vertex1_rotated = rotate_point(vertex1[0], vertex1[1], center_x, center_y, rotation_angle)
        vertex2_rotated = rotate_point(vertex2[0], vertex2[1], center_x, center_y, rotation_angle)
        vertex3_rotated = rotate_point(vertex3[0], vertex3[1], center_x, center_y, rotation_angle)

        # Draw the rotated triangle
        vertices_rotated = np.array([vertex1_rotated, vertex2_rotated, vertex3_rotated], dtype=np.int32)
        cv2.fillPoly(big_map_img, [vertices_rotated], color=(0+20*i, 255-20*i, 0))
        # 將所有中心點連接成直線
        for i in range(len(center_points) - 1):
            cv2.line(big_map_img, tuple(map(int, center_points[i])), tuple(map(int, center_points[i + 1])), (0, 0, 255), 10)

        # 儲存到資料夾
        path_file = os.path.join(folder_name, f'path_{total_path}.jpg')
        blocks = get_blocks_by_indices(all_blocks, block_indices_sorted)
        cv2.imwrite(path_file, big_map_img)
        total_path += 1



        # cv2.imshow('Merged Image', merged_image)
        # cv2.waitKey(0)

        # sorted_blocks = [blocks[index] for index in block_indices_sorted]
        blocks = get_blocks_by_indices(all_blocks, block_indices_sorted)
start_time = time.time()
clear_excel_content(excel_file)
main(blocks)

# cv2.waitKey(0)
# cv2.destroyAllWindows()

print("finish")
# 在程式開始運行前記錄時間
# 在程式執行完成後記錄時間
end_time = time.time()
# 計算程式運行的總時間
execution_time = end_time - start_time
# 將運行時間打印出來
print(f"程式運行時間： {execution_time} 秒")



# ----------------------------------------------------------------------------------------------------------


